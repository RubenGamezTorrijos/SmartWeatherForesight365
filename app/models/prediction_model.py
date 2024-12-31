import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import logging

try:
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Excel formatting will be limited.")

class PredictionModel:
    def __init__(self, model_controller):
        self.model_controller = model_controller

    def generate_prediction(self, city: str, start_date: datetime, end_date: datetime) -> pd.DataFrame:
        """
        Genera predicciones del clima y las guarda en un archivo Excel o CSV.
        """
        try:
            self.model_controller.load_models(city)

            days = (end_date - start_date).days + 1
            dates = [start_date + timedelta(days=x) for x in range(days)]
            
            # Obtener datos históricos para inicializar las predicciones
            historical_data = pd.read_csv(f"datasets/{city}_Historial_Weather.csv")
            last_row = historical_data.iloc[-1]

            # Inicializar con el último valor conocido
            initial_values = {
                'temp': float(last_row.get('tavg', last_row.get('temp', 20))),
                'temp_min': float(last_row.get('tmin', last_row.get('temp_min', 15))),
                'temp_max': float(last_row.get('tmax', last_row.get('temp_max', 25))),
                'precipitacion': float(last_row.get('prcp', 0)),
                'nieve': float(last_row.get('snow', 0)),
                'direccion_viento': float(last_row.get('wdir', 0)),
                'velocidad_viento': float(last_row.get('wspd', 0)),
                'punto_rocio': float(last_row.get('wpgt', 10)),
                'presion': float(last_row.get('pres', 1013))
            }

            predictions = {k: [v] for k, v in initial_values.items()}

            # Generar predicciones
            for _ in range(days - 1):
                current_features = np.array([predictions[k][-1] for k in predictions.keys()])
                
                for target in predictions.keys():
                    try:
                        feature_indices = [i for i, k in enumerate(predictions.keys()) if k != target]
                        features_for_target = current_features[feature_indices]
                        scaled_features = self.model_controller.scalers[target].transform([features_for_target])
                        pred = self.model_controller.models[target].predict(scaled_features)[0]
                        pred = self._apply_reasonable_limits(target, pred)
                        predictions[target].append(pred)
                    except Exception as e:
                        logging.error(f"Error al predecir {target}: {str(e)}")
                        predictions[target].append(predictions[target][-1])

            # Crear DataFrame con predicciones y nombres de columnas en español
            prediction_df = pd.DataFrame({
                'Fecha': [d.strftime('%d/%m/%Y') for d in dates],
                'Temperatura Media (°C)': predictions['temp'],
                'Temperatura Mínima (°C)': predictions['temp_min'],
                'Temperatura Máxima (°C)': predictions['temp_max'],
                'Precipitación (mm)': predictions['precipitacion'],
                'Nieve (cm)': predictions['nieve'],
                'Dirección del Viento (°)': predictions['direccion_viento'],
                'Velocidad del Viento (km/h)': predictions['velocidad_viento'],
                'Punto de Rocío (°C)': predictions['punto_rocio'],
                'Presión Atmosférica (hPa)': predictions['presion'],
                'Ciudad': city
            })

            # Reemplazar valores nulos con 'N/A'
            prediction_df = prediction_df.fillna('N/A')

            # Guardar predicciones en Excel con formato
            os.makedirs("predictions", exist_ok=True)
            excel_path = f"predictions/{city}_Future_Predictions.xlsx"
            
            if OPENPYXL_AVAILABLE:
                try:
                    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                        prediction_df.to_excel(writer, sheet_name='Predicciones', index=False)
                        
                        # Obtener la hoja de trabajo
                        workbook = writer.book
                        worksheet = writer.sheets['Predicciones']

                        # Formato para encabezados
                        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                        header_font = Font(color='FFFFFF', bold=True)
                        
                        # Aplicar formato a encabezados
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                        # Ajustar ancho de columnas
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)
                            
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column_letter].width = adjusted_width

                        # Centrar datos
                        for row in worksheet.iter_rows(min_row=2):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center')

                    logging.info(f"Archivo Excel creado exitosamente: {excel_path}")
                except Exception as e:
                    logging.error(f"Error al guardar el archivo Excel: {str(e)}")
                    raise
            else:
                # Si openpyxl no está disponible, guardar como CSV
                csv_path = f"predictions/{city}_Future_Predictions.csv"
                prediction_df.to_csv(csv_path, index=False)
                logging.info(f"Archivo CSV creado: {csv_path}")

            return prediction_df

        except Exception as e:
            logging.error(f"Error al generar predicciones: {str(e)}")
            raise

    def _apply_reasonable_limits(self, target: str, value: float) -> float:
        """
        Aplica límites razonables a las predicciones para evitar valores extremos.
        """
        limits = {
            'temp': (-10, 45),
            'temp_min': (-15, 35),
            'temp_max': (-5, 50),
            'precipitacion': (0, 500),
            'nieve': (0, 100),
            'direccion_viento': (0, 360),
            'velocidad_viento': (0, 200),
            'punto_rocio': (-20, 30),
            'presion': (900, 1100)
        }

        if target in limits:
            min_val, max_val = limits[target]
            return max(min_val, min(max_val, value))
        return value

    def _get_last_known_features(self, city: str) -> np.array:
        """Obtiene las últimas características conocidas de los datos históricos"""
        try:
            data = pd.read_csv(f"datasets/{city}_Historial_Weather.csv")
            
            column_mapping = {
                'tavg': 'temp',
                'tmin': 'temp_min',
                'tmax': 'temp_max',
                'prcp': 'precipitacion',
                'snow': 'nieve',
                'wdir': 'direccion_viento',
                'wspd': 'velocidad_viento',
                'wpgt': 'punto_rocio',
                'pres': 'presion'
            }
            
            data = data.rename(columns={col: column_mapping.get(col, col) for col in data.columns})
            
            features = ['temp', 'temp_min', 'temp_max', 'precipitacion', 'nieve', 
                       'direccion_viento', 'velocidad_viento', 'punto_rocio', 'presion']
            
            last_features = []
            for feature in features:
                if feature in data.columns:
                    value = data.iloc[-1][feature]
                    last_features.append(float(value) if not pd.isna(value) else 0)
                else:
                    last_features.append(0)
            
            return np.array(last_features)
            
        except Exception as e:
            logging.error(f"Error al obtener las últimas características para {city}: {str(e)}")
            raise

