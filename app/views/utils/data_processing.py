import streamlit as st
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def process_prediction_data(prediction_data, city):
    try:
        # Crear un archivo Excel en memoria
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Predicciones"

        # Escribir encabezados
        headers = list(prediction_data.columns)
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Escribir datos
        for row, data in enumerate(prediction_data.values, start=2):
            for col, value in enumerate(data, start=1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="center")

        # Ajustar el ancho de las columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Guardar el libro de trabajo en el objeto BytesIO
        workbook.save(output)
        output.seek(0)

        return output  # Devolver el archivo Excel como un objeto en memoria

    except Exception as e:
        st.error(f"Error al preparar el archivo para descarga: {str(e)}")
        return None